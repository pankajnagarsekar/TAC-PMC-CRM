"""
Export Service - Template-driven Excel and PDF export engine
Handles report generation with strict compliance to spec templates
"""

from typing import Dict, Any, List, Optional
from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
import json
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from weasyprint import HTML, CSS
    HAS_WEASYPRINT = True
except Exception:
    # WeasyPrint may raise OSError when native GTK/Pango libs are missing.
    # Keep API startup alive and fail only when PDF export is requested.
    HAS_WEASYPRINT = False


class ExportService:
    """
    Template-driven export engine for Excel and PDF reports.
    Maintains exact field layout and formula protection per spec.
    """

    # Template configurations for each report type
    REPORT_TEMPLATES = {
        "project_summary": {
            "description": "Project-level financial summary",
            "excel_sheet": "Project Summary",
            "template_file": "project_summary.xlsx",
            "header_row": 7,
            "columns": [
                ("CODE", 10),
                ("Description", 30),
                ("WO Value", 15),
                ("% Progress", 12),
                ("Payment Value", 15),
                ("Deadline", 15),
                ("Difference", 15),
            ]
        },
        "work_order_tracker": {
            "description": "Work Order tracking report",
            "excel_sheet": "WO Tracker",
            "template_file": "work_order_tracker.xlsx",
            "header_row": 6,
            "columns": [
                ("CODE", 10),
                ("WO Reference", 20),
                ("Vendor", 20),
                ("WO Value", 15),
                ("Retention Value", 15),
                ("Start Date", 15),
                ("End Date", 15),
            ]
        },
        "payment_certificate_tracker": {
            "description": "Payment Certificate tracking report",
            "excel_sheet": "PC Tracker",
            "template_file": "payment_certificate_tracker.xlsx",
            "header_row": 6,
            "columns": [
                ("CODE", 10),
                ("PC Reference", 20),
                ("Vendor", 20),
                ("PC Value", 15),
                ("PC Date", 15),
                ("Payment Value", 15),
                ("Payment Date", 15),
            ]
        },
        "petty_cash_tracker": {
            "description": "Petty Cash and OVH transaction report",
            "excel_sheet": "Petty Cash",
            "template_file": "petty_cash_tracker.xlsx",
            "header_row": 6,
            "columns": [
                ("Date", 15),
                ("PC Refn", 20),
                ("PC Value", 15),
                ("Bill / Invoice", 30),
            ]
        },
        "csa_report": {
            "description": "Category-Specific Activity report",
            "excel_sheet": "CSA Tracker",
            "template_file": "csa_tracker.xlsx",
            "header_row": 6,
            "columns": [
                ("CODE", 10),
                ("WO Refn", 20),
                ("Description", 30),
                ("Qty", 10),
                ("Received Date", 15),
            ]
        },
        "weekly_progress": {
            "description": "Weekly progress report",
            "excel_sheet": "Weekly Progress",
            "template_file": "weekly_progress.xlsx",
            "header_row": 6,
            "columns": [
                ("CODE", 10),
                ("WO Refn", 20),
                ("Vendor", 20),
                ("% Completed", 15),
                ("Comments", 30),
            ]
        },
        "15_days_progress": {
            "description": "15-day progress report",
            "excel_sheet": "Weekly Progress", # Sharing template for now or use same structure
            "template_file": "weekly_progress.xlsx",
            "header_row": 6,
            "columns": [
                ("CODE", 10),
                ("WO Refn", 20),
                ("Vendor", 20),
                ("% Completed", 15),
                ("Comments", 30),
            ]
        },
        "monthly_progress": {
            "description": "Monthly progress report",
            "excel_sheet": "Weekly Progress",
            "template_file": "weekly_progress.xlsx",
            "header_row": 6,
            "columns": [
                ("CODE", 10),
                ("WO Refn", 20),
                ("Vendor", 20),
                ("% Completed", 15),
                ("Comments", 30),
            ]
        },
    }

    @staticmethod
    def format_currency(value: Decimal | float | int) -> str:
        """Format value as Indian currency (₹ X,XX,XXX.XX)"""
        if isinstance(value, str):
            try:
                value = Decimal(value)
            except:
                return value
        
        if isinstance(value, Decimal):
            value = float(value)
        
        # Format with thousands separator and 2 decimals
        return f"₹{value:,.2f}"

    @staticmethod
    def export_to_excel(
        report_type: str,
        report_data: Dict[str, Any],
        company_info: Optional[Dict[str, Any]] = None,
        include_terms: bool = True,
        terms_text: Optional[str] = None
    ) -> bytes:
        """
        Generate Excel export from structured report data.
        Maintains template structure with data injection.
        """
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed. Install with: pip install openpyxl")

        if report_type not in ExportService.REPORT_TEMPLATES:
            raise ValueError(f"Unknown report type: {report_type}")

        template = ExportService.REPORT_TEMPLATES[report_type]
        template_filename = template.get("template_file")
        header_row = template.get("header_row", 1)
        
        # Load from template if available, else create new
        template_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "templates")
        template_path = os.path.join(template_dir, template_filename) if template_filename else None
        
        if template_path and os.path.exists(template_path):
            from openpyxl import load_workbook
            wb = load_workbook(template_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = template["excel_sheet"]

        # Styling
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        currency_format = "_-₹* #,##0.00_-;-₹* #,##0.00_-;_-₹* \"-\"??_-;_-@_-"
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Add header row (if not using pre-templated headers, or to ensure they are set)
        for col_idx, (col_name, col_width) in enumerate(template["columns"], 2): # Start from col B (2)
            cell = ws.cell(row=header_row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = col_width

        # Add data rows from report_data
        rows = report_data.get("rows", [])
        for row_idx, row_data in enumerate(rows, header_row + 1):
            for col_idx, cell_value in enumerate(row_data, 2): # Start from col B (2)
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = border
                
                # Right-align numeric columns
                if isinstance(cell_value, (int, float, Decimal)):
                    cell.alignment = Alignment(horizontal="right")
                    # Check if this column should be formatted as currency
                    col_header = template["columns"][col_idx - 2][0].lower()
                    if any(term in col_header for term in ["amount", "budget", "value", "payable", "certified"]):
                        cell.number_format = currency_format
                else:
                    cell.alignment = Alignment(horizontal="left")

        # Inject Project Info into Template Fields (B2, B3, B4 usually)
        if company_info:
            ws['C2'] = company_info.get("project_name", "")
            ws['C3'] = company_info.get("client_name", "")
            ws['C4'] = datetime.now(timezone.utc).strftime("%Y-%m-%d")

        # Add Terms & Conditions if requested
        if include_terms and terms_text:
            if "Terms & Conditions" in wb.sheetnames:
                terms_ws = wb["Terms & Conditions"]
            else:
                terms_ws = wb.create_sheet("Terms & Conditions")
            
            terms_ws["A1"] = "TERMS & CONDITIONS"
            terms_ws["A2"] = terms_text
            terms_ws["A2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            terms_ws.column_dimensions["A"].width = 100

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @staticmethod
    def export_to_pdf(
        report_type: str,
        report_data: Dict[str, Any],
        company_info: Optional[Dict[str, Any]] = None,
        include_terms: bool = True,
        terms_text: Optional[str] = None
    ) -> bytes:
        """
        Generate PDF export using WeasyPrint HTML->PDF conversion.
        Mirrors Excel layout exactly using CSS styling.
        
        Args:
            report_type: Type of report
            report_data: Dictionary containing rows and metadata
            company_info: Company branding info
            include_terms: Whether to append T&C
            terms_text: Full T&C text
            
        Returns:
            PDF file content as bytes
        """
        if not HAS_WEASYPRINT:
            raise RuntimeError("weasyprint not installed. Install with: pip install weasyprint")

        if report_type not in ExportService.REPORT_TEMPLATES:
            raise ValueError(f"Unknown report type: {report_type}")

        template = ExportService.REPORT_TEMPLATES[report_type]
        
        # Build HTML from report data
        html_content = ExportService._build_html_report(
            report_type, report_data, company_info, include_terms, terms_text
        )

        # CSS for PDF styling
        css_string = """
        @page {
            size: A4;
            margin: 20mm;
        }
        body {
            font-family: 'Segoe UI', Arial, sans-serif;
            line-height: 1.5;
            color: #333;
        }
        .header {
            text-align: center;
            margin-bottom: 20px;
            border-bottom: 2px solid #1E3A5F;
            padding-bottom: 10px;
        }
        .logo {
            height: 60px;
            margin-bottom: 10px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-bottom: 20px;
        }
        th {
            background-color: #1E3A5F;
            color: white;
            padding: 10px;
            text-align: left;
            border: 1px solid #ddd;
        }
        td {
            padding: 8px;
            border: 1px solid #ddd;
        }
        tr:nth-child(even) {
            background-color: #f9f9f9;
        }
        .currency {
            text-align: right;
            font-family: 'Courier New', monospace;
        }
        .terms {
            page-break-before: always;
            font-size: 9pt;
            margin-top: 20px;
        }
        """

        # Convert HTML to PDF
        html = HTML(string=html_content, base_url=".")
        css = CSS(string=css_string)
        pdf_bytes = html.write_pdf(stylesheets=[css])

        return pdf_bytes

    @staticmethod
    def _build_html_report(
        report_type: str,
        report_data: Dict[str, Any],
        company_info: Optional[Dict[str, Any]],
        include_terms: bool,
        terms_text: Optional[str]
    ) -> str:
        """Build HTML string for PDF generation"""
        template = ExportService.REPORT_TEMPLATES[report_type]
        
        html_parts = [
            "<html><head><meta charset='UTF-8'></head><body>",
            "<div class='header'>",
        ]

        # Add company info if available
        if company_info:
            html_parts.append(f"<h1>{company_info.get('name', 'Report')}</h1>")
            html_parts.append(f"<p>{company_info.get('address', '')}</p>")

        html_parts.append(f"<h2>{template['description']}</h2>")
        html_parts.append(f"<p>Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}</p>")
        html_parts.append("</div>")

        # Add table
        html_parts.append("<table>")
        html_parts.append("<thead><tr>")
        for col_name, _ in template["columns"]:
            html_parts.append(f"<th>{col_name}</th>")
        html_parts.append("</tr></thead>")

        html_parts.append("<tbody>")
        rows = report_data.get("rows", [])
        for row_data in rows:
            html_parts.append("<tr>")
            for col_idx, cell_value in enumerate(row_data):
                is_currency = "Amount" in template["columns"][col_idx][0] or "Budget" in template["columns"][col_idx][0]
                css_class = "currency" if is_currency and isinstance(cell_value, (int, float, Decimal)) else ""
                html_parts.append(f"<td class='{css_class}'>{cell_value}</td>")
            html_parts.append("</tr>")
        html_parts.append("</tbody>")
        html_parts.append("</table>")

        # Add terms if requested
        if include_terms and terms_text:
            html_parts.append("<div class='terms'>")
            html_parts.append("<h3>TERMS & CONDITIONS</h3>")
            html_parts.append(f"<p>{terms_text}</p>")
            html_parts.append("</div>")

        html_parts.append("</body></html>")

        return "".join(html_parts)

    @staticmethod
    def validate_report_type(report_type: str) -> bool:
        """Validate if report type is supported"""
        return report_type in ExportService.REPORT_TEMPLATES

    @staticmethod
    def list_supported_reports() -> List[Dict[str, Any]]:
        """List all supported report types"""
        return [
            {
                "type": report_type,
                "description": template["description"],
                "excel_sheet": template["excel_sheet"],
            }
            for report_type, template in ExportService.REPORT_TEMPLATES.items()
        ]
