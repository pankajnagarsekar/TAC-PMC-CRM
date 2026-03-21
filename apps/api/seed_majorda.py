#!/usr/bin/env python3
"""
Majorda Villa — Deterministic Database Seed Script
====================================================
Source: apps/web/memory/majorda templates.xlsx
Schema: Backend Database Schema & Financial Integrity Specification v3.0

Usage:
    python seed_majorda.py                          # seed
    python seed_majorda.py --cleanup                # cleanup only
    python seed_majorda.py --cleanup --seed         # full reset + seed

Re-runnable: cleanup_db() deletes all records scoped to SEED_ORG_NAME.
"""

import os
import sys
import io
import argparse

# Force UTF-8 output on Windows (avoids cp1252 UnicodeEncodeError)
if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from datetime import datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

import openpyxl
from bson import ObjectId, Decimal128
from passlib.context import CryptContext
from pymongo import MongoClient, ASCENDING, DESCENDING

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
MONGO_URL = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
DB_NAME   = os.environ.get("DB_NAME",   "tac_pmc_crm")

EXCEL_PATH = Path(__file__).parent.parent / "web" / "memory" / "majorda templates.xlsx"

SEED_ORG_NAME     = "Third Angle Concepts (PMC)"
SEED_PROJECT_NAME = "Majorda Villa"
SEED_CLIENT_NAME  = "Mr. Sanjay Rao"
SEED_VENDOR_NAME  = "ABC Contractors"
DEFAULT_PASSWORD  = "Admin@1234"

CGST_PCT = Decimal("0")   # WO Tracker has no GST values; treat as zero-GST WOs
SGST_PCT = Decimal("0")   # (all financial values from sheet are treated as grand_total)
RETENTION_PCT = Decimal("5")

pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def now() -> datetime:
    return datetime.now(timezone.utc)

def d128(value) -> Decimal128:
    """Convert numeric/Decimal to Decimal128 for MongoDB storage."""
    return Decimal128(str(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)))

def pct(value) -> Decimal128:
    """Store a percentage as Decimal128."""
    return Decimal128(str(Decimal(str(value)).quantize(Decimal("0.0000"), rounding=ROUND_HALF_UP)))

def oid() -> ObjectId:
    return ObjectId()

def warn(msg: str):
    print(f"  [WARN] {msg}")

def ok(msg: str):
    print(f"  [OK]   {msg}")

def section(title: str):
    print(f"\n{'='*60}")
    print(f"  {title}")
    print(f"{'='*60}")

# ---------------------------------------------------------------------------
# Excel Parsing
# ---------------------------------------------------------------------------
def load_excel() -> dict:
    """
    Parse all relevant sheets from the Excel workbook.
    Returns a dict with keys: categories, project_summary, wo_tracker,
    pc_tracker, weekly_progress.
    """
    print(f"\nLoading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    # ── Code-Categories ─────────────────────────────────────────────────────
    ws_cat = wb["Code-Categories"]
    categories = []
    for row in ws_cat.iter_rows(min_row=3, values_only=True):  # skip header rows
        if row[1] is None:
            continue
        categories.append({
            "code":        str(row[1]).strip(),
            "name":        str(row[2]).strip() if row[2] else "",
            "inclusion":   str(row[3]).strip() if row[3] else "",
            # OVH is fund_transfer; CSA is fund_transfer; rest are commitment
            "budget_type": "fund_transfer" if str(row[1]).strip() in ("OVH", "CSA") else "commitment",
        })
    ok(f"Loaded {len(categories)} categories")

    # ── Project Summary (budget per code) ───────────────────────────────────
    ws_ps = wb["Project Summary"]
    project_summary = {}
    for row in ws_ps.iter_rows(min_row=8, values_only=True):  # skip header block
        if row[1] is None or row[2] is None:
            continue
        code = str(row[1]).strip()
        if code == "Total:":
            break
        project_summary[code] = {
            "wo_value":      Decimal(str(row[3])) if row[3] else Decimal("0"),
            "pct_progress":  Decimal(str(row[4])) if row[4] else Decimal("0"),
            "payment_value": Decimal(str(row[5])) if row[5] else Decimal("0"),
        }
    ok(f"Loaded project summary for {len(project_summary)} codes")

    # ── Work Order Tracker ───────────────────────────────────────────────────
    ws_wo = wb["Work Order Tracker"]
    wo_tracker = []
    for row in ws_wo.iter_rows(min_row=7, values_only=True):  # skip header block
        if row[1] is None or row[4] is None:
            continue
        code = str(row[1]).strip()
        if code == "Total:":
            break
        wo_value      = Decimal(str(row[4])) if row[4] else Decimal("0")
        retention_val = Decimal(str(row[5])) if row[5] else Decimal("0")
        expected_ret  = (wo_value * RETENTION_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        if retention_val != expected_ret:
            warn(f"WO {code}: Retention in sheet = {retention_val}, "
                 f"expected {RETENTION_PCT}% of {wo_value} = {expected_ret}. "
                 f"Proceeding with sheet value.")
        wo_tracker.append({
            "code":          code,
            "wo_ref":        str(row[2]).strip() if row[2] else f"TAC_WO_25_{code}",
            "vendor_name":   str(row[3]).strip() if row[3] else SEED_VENDOR_NAME,
            "wo_value":      wo_value,
            "retention_val": retention_val,  # from sheet (validated above)
            "start_date":    row[6],
            "end_date":      row[7],
        })
    ok(f"Loaded {len(wo_tracker)} work order rows")

    # ── Payment Certificate Tracker ──────────────────────────────────────────
    ws_pc = wb["Payment Certificate Tracker"]
    pc_tracker = []
    for row in ws_pc.iter_rows(min_row=7, values_only=True):
        if row[1] is None or row[4] is None:
            continue
        code = str(row[1]).strip()
        if code == "Total:":
            break
        pc_value      = Decimal(str(row[4])) if row[4] else Decimal("0")
        payment_value = Decimal(str(row[6])) if row[6] else Decimal("0")
        # Spec: retention_amount = subtotal * retention_pct
        # PC Value in sheet = subtotal (before retention)
        expected_ret = (pc_value * RETENTION_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        net_payable   = pc_value - expected_ret
        # Sanity: net_payable (PC Value - 5% retention) should equal Payment Value.
        # NOTE: The Excel sheet's "Payment Value" column records the GROSS PC Value
        # (before retention deduction), so this warning is EXPECTED for all rows in
        # the Majorda template — the spec formula is authoritative.
        if payment_value and abs(payment_value - net_payable) > Decimal("1"):
            warn(f"PC {code}: Excel Payment Value {payment_value} != PC Value - 5% retention "
                 f"({net_payable}). The sheet stores gross value; spec formula applied.")
        pc_tracker.append({
            "code":          code,
            "pc_ref":        str(row[2]).strip() if row[2] else f"TAC_PC_25_{code}",
            "vendor_name":   str(row[3]).strip() if row[3] else SEED_VENDOR_NAME,
            "pc_value":      pc_value,         # subtotal
            "retention_val": expected_ret,
            "payment_value": payment_value,
            "pc_date":       row[5],
        })
    ok(f"Loaded {len(pc_tracker)} payment certificate rows")

    # ── Weekly Progress Report (DPR source) ──────────────────────────────────
    ws_wp = wb["Weekly Progress Report"]
    weekly_progress = {}
    for row in ws_wp.iter_rows(min_row=7, values_only=True):
        if row[1] is None:
            continue
        code = str(row[1]).strip()
        if code == "Total:":
            break
        pct_complete = Decimal(str(row[4])) * 100 if row[4] else Decimal("0")
        weekly_progress[code] = {
            "wo_ref":       str(row[2]).strip() if row[2] else "",
            "vendor":       str(row[3]).strip() if row[3] else "",
            "pct_complete": pct_complete,
            "comments":     str(row[5]).strip() if row[5] else "",
        }
    ok(f"Loaded weekly progress for {len(weekly_progress)} codes")

    return {
        "categories":      categories,
        "project_summary": project_summary,
        "wo_tracker":      wo_tracker,
        "pc_tracker":      pc_tracker,
        "weekly_progress": weekly_progress,
    }


# ---------------------------------------------------------------------------
# PC ↔ WO Financial Validation
# ---------------------------------------------------------------------------
def validate_pc_vs_wo(wo_tracker: list, pc_tracker: list):
    """
    Per spec: PC subtotal + retention must not exceed WO grand_total.
    The sum of all PCs per code should be <= WO value.
    """
    section("Financial Validation: PC vs WO")
    wo_map = {row["code"]: row for row in wo_tracker}
    pc_by_code: dict = {}
    for pc in pc_tracker:
        pc_by_code.setdefault(pc["code"], []).append(pc)

    for code, pcs in pc_by_code.items():
        if code not in wo_map:
            warn(f"PC code '{code}' has no matching WO. Skipping validation.")
            continue
        wo_val      = wo_map[code]["wo_value"]
        total_pc    = sum(p["pc_value"] for p in pcs)
        total_ret   = sum(p["retention_val"] for p in pcs)
        total_gross = total_pc + total_ret
        if total_gross > wo_val:
            warn(f"Code {code}: SUM(PC value + retention) = {total_gross} "
                 f"> WO value {wo_val}. Over-certification!")
        else:
            ok(f"Code {code}: SUM(PC+retention)={total_gross} <= WO={wo_val} -- OK")


# ---------------------------------------------------------------------------
# Cleanup
# ---------------------------------------------------------------------------
def cleanup_db(db) -> None:
    section("Cleanup — removing seed data")

    org = db.organisations.find_one({"name": SEED_ORG_NAME})
    if not org:
        print("  No seed organisation found — nothing to clean.")
        return

    org_id = str(org["_id"])
    project = db.projects.find_one({"organisation_id": org_id,
                                    "project_name": SEED_PROJECT_NAME})
    project_id = str(project["_id"]) if project else None

    collections_org_scoped = [
        "users", "clients", "vendors", "code_master",
        "global_settings", "notifications",
    ]
    collections_project_scoped = [
        "projects", "project_category_budgets", "work_orders",
        "payment_certificates", "financial_state", "fund_allocations",
        "cash_transactions", "vendor_ledger", "site_overheads",
        "dprs", "attendance", "workers_daily_logs", "voice_logs",
        "audit_logs",
    ]

    for col in collections_org_scoped:
        r = db[col].delete_many({"organisation_id": org_id})
        if r.deleted_count:
            print(f"  Deleted {r.deleted_count:>4d} from {col}")

    if project_id:
        for col in collections_project_scoped:
            r = db[col].delete_many({"project_id": project_id})
            if r.deleted_count:
                print(f"  Deleted {r.deleted_count:>4d} from {col}")

    # Delete project itself
    db.projects.delete_many({"organisation_id": org_id})
    # Delete sequences
    db.sequences.delete_many({"_id": {"$regex": f".*_{org_id}"}})
    # user_project_map
    db.user_project_map.delete_many({"project_id": project_id} if project_id else {})
    # Delete org last
    db.organisations.delete_one({"_id": org["_id"]})
    ok("Cleanup complete")


# ---------------------------------------------------------------------------
# Seed
# ---------------------------------------------------------------------------
def seed(db, data: dict) -> None:
    cats      = data["categories"]
    ps        = data["project_summary"]
    wo_rows   = data["wo_tracker"]
    pc_rows   = data["pc_tracker"]
    wp        = data["weekly_progress"]

    # ── 1. Organisation ─────────────────────────────────────────────────────
    section("1. Organisation")
    org_doc = {
        "_id":        oid(),
        "name":       SEED_ORG_NAME,
        "created_at": now(),
    }
    db.organisations.insert_one(org_doc)
    org_id = str(org_doc["_id"])
    ok(f"Organisation: {SEED_ORG_NAME}  id={org_id}")

    # ── 2. Global Settings ──────────────────────────────────────────────────
    section("2. Global Settings")
    gs_doc = {
        "_id":                 oid(),
        "organisation_id":     org_id,
        "name":                SEED_ORG_NAME,
        "address":             "DSR Elite, No 109, Mahadevpura, Bengaluru, Karnataka 560048",
        "email":               "info@thirdangleconcepts.com",
        "phone":               "7387471134",
        "gst_number":          "",
        "pan_number":          "",
        "cgst_percentage":     d128(9),
        "sgst_percentage":     d128(9),
        "retention_percentage": d128(5),
        "wo_prefix":           "TAC_WO",
        "pc_prefix":           "TAC_PC",
        "invoice_prefix":      "TAC_INV",
        "terms_and_conditions": (
            "5% retention applicable on all PCs (Ref 16). "
            "Payment Certificates prepared between 25th–30th of each month."
        ),
        "currency":        "INR",
        "currency_symbol": "₹",
        "client_permissions": {
            "can_view_dpr":        True,
            "can_view_financials": False,
            "can_view_reports":    True,
        },
        "updated_at": now(),
    }
    db.global_settings.insert_one(gs_doc)
    ok("Global settings inserted")

    # ── 3. Users ────────────────────────────────────────────────────────────
    section("3. Users (Admin + Supervisor)")
    admin_id = oid()
    super_id = oid()
    admin_doc = {
        "_id":                      admin_id,
        "organisation_id":          org_id,
        "name":                     "Amit Chinchuwar",
        "email":                    "amit@thirdangleconcepts.com",
        "hashed_password":          pwd_ctx.hash(DEFAULT_PASSWORD),
        "role":                     "Admin",
        "active_status":            True,
        "dpr_generation_permission": True,
        "assigned_projects":        [],
        "screen_permissions":       [],
        "created_at":               now(),
        "updated_at":               now(),
    }
    supervisor_doc = {
        "_id":                      super_id,
        "organisation_id":          org_id,
        "name":                     "Site Supervisor",
        "email":                    "supervisor@thirdangleconcepts.com",
        "hashed_password":          pwd_ctx.hash(DEFAULT_PASSWORD),
        "role":                     "Supervisor",
        "active_status":            True,
        "dpr_generation_permission": True,
        "assigned_projects":        [],
        "screen_permissions":       [],
        "created_at":               now(),
        "updated_at":               now(),
    }
    db.users.insert_many([admin_doc, supervisor_doc])
    ok(f"Admin:      {admin_doc['email']}")
    ok(f"Supervisor: {supervisor_doc['email']}")
    ok(f"Password:   {DEFAULT_PASSWORD}")

    # ── 4. Client ───────────────────────────────────────────────────────────
    section("4. Client")
    client_doc = {
        "_id":             oid(),
        "organisation_id": org_id,
        "name":            SEED_CLIENT_NAME,
        "address":         "Plot No-51/16, Beach Road, Panchecho Waddo, Betalbatim, Majorda, Goa 403713",
        "phone":           None,
        "email":           None,
        "gstin":           None,
        "created_at":      now(),
        "updated_at":      now(),
    }
    db.clients.insert_one(client_doc)
    client_id = str(client_doc["_id"])
    ok(f"Client: {SEED_CLIENT_NAME}  id={client_id}")

    # ── 5. Sequences (WO + PC) ──────────────────────────────────────────────
    section("5. Sequences")
    db.sequences.insert_many([
        {"_id": f"wo_seq_{org_id}", "seq": 0},
        {"_id": f"pc_seq_{org_id}", "seq": 0},
    ])
    ok("Sequences initialised: wo_seq=0, pc_seq=0")

    # ── 6. Code Master (14 categories) ──────────────────────────────────────
    section("6. Code Master (categories)")
    cat_id_map: dict[str, str] = {}  # code → str(_id)
    cat_docs = []
    for cat in cats:
        doc = {
            "_id":              oid(),
            "organisation_id":  org_id,
            "category_name":    cat["name"],
            "code":             cat["code"],
            "code_short":       cat["code"],
            "code_description": cat["name"],
            "budget_type":      cat["budget_type"],
            "active_status":    True,
            "created_at":       now(),
            "updated_at":       now(),
        }
        cat_docs.append(doc)
        cat_id_map[cat["code"]] = str(doc["_id"])
    db.code_master.insert_many(cat_docs)
    ok(f"Inserted {len(cat_docs)} categories")
    for c in cats:
        print(f"    {c['code']:5s} → {c['name']:<45s} [{c['budget_type']}]")

    # ── 7. Vendor ───────────────────────────────────────────────────────────
    section("7. Vendor")
    vendor_doc = {
        "_id":             oid(),
        "organisation_id": org_id,
        "name":            SEED_VENDOR_NAME,
        "gstin":           None,
        "contact_person":  "Mr. ABC",
        "phone":           None,
        "email":           None,
        "address":         None,
        "active_status":   True,
        "created_at":      now(),
        "updated_at":      now(),
    }
    db.vendors.insert_one(vendor_doc)
    vendor_id = str(vendor_doc["_id"])
    ok(f"Vendor: {SEED_VENDOR_NAME}  id={vendor_id}")

    # ── 8. Project ──────────────────────────────────────────────────────────
    section("8. Project")
    total_budget = sum(
        Decimal(str(row["wo_value"])) for row in wo_rows
    )
    project_doc = {
        "_id":                          oid(),
        "organisation_id":              org_id,
        "project_name":                 SEED_PROJECT_NAME,
        "client_id":                    client_id,
        "project_code":                 "MAJ-25",
        "status":                       "active",
        "address":                      "Plot No-51/16, Beach Road, Panchecho Waddo, Betalbatim",
        "city":                         "Majorda",
        "state":                        "Goa",
        "project_retention_percentage": pct(5),
        "project_cgst_percentage":      pct(9),
        "project_sgst_percentage":      pct(9),
        "completion_percentage":        pct(0),
        "master_original_budget":       d128(total_budget),
        "master_remaining_budget":      d128(total_budget),
        "threshold_petty":              d128(10000),
        "threshold_ovh":                d128(50000),
        "version":                      1,
        "created_at":                   now(),
        "updated_at":                   now(),
    }
    db.projects.insert_one(project_doc)
    project_id = str(project_doc["_id"])
    ok(f"Project: {SEED_PROJECT_NAME}  id={project_id}")
    ok(f"Total master budget: ₹{total_budget:,.2f}")

    # Update users' assigned_projects
    db.users.update_many(
        {"organisation_id": org_id},
        {"$set": {"assigned_projects": [project_id]}}
    )
    # user_project_map
    db.user_project_map.insert_many([
        {"_id": oid(), "user_id": str(admin_id),
         "project_id": project_id, "created_at": now()},
        {"_id": oid(), "user_id": str(super_id),
         "project_id": project_id, "created_at": now()},
    ])
    ok("user_project_map entries created")

    # ── 9. Project Category Budgets ─────────────────────────────────────────
    section("9. Project Category Budgets")
    budget_docs = []
    budget_map: dict[str, str] = {}  # category_id → budget doc _id (for reference)
    for row in wo_rows:
        code = row["code"]
        cat_id = cat_id_map.get(code)
        if not cat_id:
            warn(f"Code '{code}' from WO tracker not found in code_master. Skipping.")
            continue
        wo_val = row["wo_value"]
        doc = {
            "_id":              oid(),
            "project_id":       project_id,
            "category_id":      cat_id,
            "original_budget":  d128(wo_val),
            "committed_amount": d128(0),
            "remaining_budget": d128(wo_val),
            "description":      None,
            "version":          1,
            "created_at":       now(),
            "updated_at":       now(),
        }
        budget_docs.append(doc)
        budget_map[cat_id] = str(doc["_id"])
    db.project_category_budgets.insert_many(budget_docs)
    ok(f"Inserted {len(budget_docs)} project_category_budgets")

    # ── 10. Work Orders + Budget commitment ─────────────────────────────────
    section("10. Work Orders")
    wo_id_map: dict[str, str] = {}   # code → wo _id string
    wo_docs = []
    ledger_docs = []
    wo_seq = 0

    for row in wo_rows:
        code   = row["code"]
        cat_id = cat_id_map.get(code)
        if not cat_id:
            warn(f"Skipping WO for unknown code: {code}")
            continue

        wo_seq += 1
        wo_value = row["wo_value"]

        # Financial calculation (WO Value = grand_total, no GST in this project)
        subtotal         = wo_value
        discount         = Decimal("0")
        total_before_tax = subtotal - discount
        cgst_amount      = (total_before_tax * CGST_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        sgst_amount      = (total_before_tax * SGST_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        grand_total      = total_before_tax + cgst_amount + sgst_amount
        retention_amount = (grand_total * RETENTION_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        total_payable    = grand_total - retention_amount
        actual_payable   = total_payable

        # Validate retention against sheet value
        sheet_ret = row["retention_val"]
        if abs(sheet_ret - retention_amount) > Decimal("1"):
            warn(f"WO {code}: Recomputed retention {retention_amount} ≠ sheet {sheet_ret}. "
                 f"Using spec formula ({RETENTION_PCT}% of grand_total).")

        wo_id = oid()
        wo_ref = row["wo_ref"] if row["wo_ref"] and "TAC_WO" in row["wo_ref"] \
                 else f"TAC_WO_25_{wo_seq:03d}"

        wo_doc = {
            "_id":               wo_id,
            "organisation_id":   org_id,
            "project_id":        project_id,
            "category_id":       cat_id,
            "vendor_id":         vendor_id,
            "wo_ref":            wo_ref,
            "subtotal":          d128(subtotal),
            "discount":          d128(discount),
            "total_before_tax":  d128(total_before_tax),
            "cgst":              d128(cgst_amount),
            "sgst":              d128(sgst_amount),
            "grand_total":       d128(grand_total),
            "retention_percent": pct(RETENTION_PCT),
            "retention_amount":  d128(retention_amount),
            "total_payable":     d128(total_payable),
            "actual_payable":    d128(actual_payable),
            "status":            "Pending",
            "line_items": [{
                "sr_no":       1,
                "description": f"{code} — {next((c['name'] for c in cats if c['code'] == code), code)} (lump sum)",
                "qty":         d128(1),
                "rate":        d128(subtotal),
                "total":       d128(subtotal),
            }],
            "version":    1,
            "created_at": now(),
            "updated_at": now(),
        }
        wo_docs.append(wo_doc)
        wo_id_map[code] = str(wo_id)

        # Vendor ledger: PC_CERTIFIED entry per WO (commitment phase)
        ledger_docs.append({
            "_id":        oid(),
            "vendor_id":  vendor_id,
            "project_id": project_id,
            "ref_id":     str(wo_id),
            "entry_type": "PC_CERTIFIED",
            "amount":     d128(grand_total),
            "created_at": now(),
        })

    db.work_orders.insert_many(wo_docs)
    ok(f"Inserted {len(wo_docs)} work orders")
    # Fix remaining_budget with proper Decimal128 (pymongo $inc with Decimal128 is tricky)
    for row in wo_rows:
        code   = row["code"]
        cat_id = cat_id_map.get(code)
        if not cat_id:
            continue
        wo_val = row["wo_value"]
        db.project_category_budgets.update_one(
            {"project_id": project_id, "category_id": cat_id},
            {"$set": {
                "committed_amount": d128(wo_val),
                "remaining_budget": d128(Decimal("0")),
                "updated_at":       now(),
            }}
        )

    # ── 11. Vendor Ledger (WO commitments) ──────────────────────────────────
    section("11. Vendor Ledger — WO commitment entries")
    db.vendor_ledger.insert_many(ledger_docs)
    ok(f"Inserted {len(ledger_docs)} vendor_ledger entries (PC_CERTIFIED / WO phase)")

    # ── 12. Payment Certificates ─────────────────────────────────────────────
    section("12. Payment Certificates")
    pc_docs        = []
    pc_ledger_docs = []
    pc_seq = 0

    for row in pc_rows:
        code   = row["code"]
        cat_id = cat_id_map.get(code)
        wo_id_str = wo_id_map.get(code)
        if not cat_id:
            warn(f"Skipping PC for unknown code: {code}")
            continue

        pc_seq += 1
        pc_value = row["pc_value"]  # subtotal (before retention)

        # Financial calculation per spec
        retention_amount = (pc_value * RETENTION_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        total_payable    = pc_value - retention_amount
        cgst_amount      = (total_payable * CGST_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        sgst_amount      = (total_payable * SGST_PCT / 100).quantize(Decimal("0.01"), ROUND_HALF_UP)
        gst_amount       = cgst_amount + sgst_amount
        grand_total      = total_payable + gst_amount

        # Validate against sheet
        sheet_ret = row["retention_val"]
        if abs(sheet_ret - retention_amount) > Decimal("1"):
            warn(f"PC {code}: Recomputed retention {retention_amount} ≠ sheet {sheet_ret}. "
                 f"Using spec formula.")

        # Validate: PC gross (subtotal+retention) must not exceed WO value
        wo_row = next((w for w in wo_rows if w["code"] == code), None)
        if wo_row:
            pc_gross = pc_value + retention_amount
            if pc_gross > wo_row["wo_value"]:
                warn(f"PC {code}: PC subtotal+retention ({pc_gross}) > WO value ({wo_row['wo_value']})")

        pc_ref = row["pc_ref"] if row["pc_ref"] and "TAC_PC" in row["pc_ref"] \
                 else f"TAC_PC_25_{pc_seq:03d}"

        pc_id = oid()
        pc_doc = {
            "_id":               pc_id,
            "organisation_id":   org_id,
            "project_id":        project_id,
            "work_order_id":     wo_id_str,
            "category_id":       cat_id,
            "vendor_id":         vendor_id,
            "pc_ref":            pc_ref,
            "subtotal":          d128(pc_value),
            "retention_percent": pct(RETENTION_PCT),
            "retention_amount":  d128(retention_amount),
            "total_payable":     d128(total_payable),
            "cgst":              d128(cgst_amount),
            "sgst":              d128(sgst_amount),
            "gst_amount":        d128(gst_amount),
            "grand_total":       d128(grand_total),
            "fund_request":      False,
            "status":            "Closed",
            "line_items": [{
                "sr_no":         1,
                "scope_of_work": f"{code} — partial billing cycle 1",
                "rate":          d128(pc_value),
                "qty":           d128(1),
                "unit":          "LS",
                "total":         d128(pc_value),
            }],
            "idempotency_key": f"seed-pc-{code}-001",
            "version":         1,
            "vendor_name":     SEED_VENDOR_NAME,
            "invoice_number":  None,
            "date":            None,
            "amount":          d128(pc_value),
            "total_amount":    d128(grand_total),
            "ocr_id":          None,
            "created_at":      now(),
        }
        pc_docs.append(pc_doc)

        # Vendor ledger: RETENTION_HELD per PC
        if retention_amount > Decimal("0"):
            pc_ledger_docs.append({
                "_id":        oid(),
                "vendor_id":  vendor_id,
                "project_id": project_id,
                "ref_id":     str(pc_id),
                "entry_type": "RETENTION_HELD",
                "amount":     d128(retention_amount),
                "created_at": now(),
            })
        # Vendor ledger: PAYMENT_MADE per PC
        pc_ledger_docs.append({
            "_id":        oid(),
            "vendor_id":  vendor_id,
            "project_id": project_id,
            "ref_id":     str(pc_id),
            "entry_type": "PAYMENT_MADE",
            "amount":     d128(total_payable),
            "created_at": now(),
        })

    db.payment_certificates.insert_many(pc_docs)
    ok(f"Inserted {len(pc_docs)} payment certificates")

    # ── 13. Vendor Ledger (PC entries) ──────────────────────────────────────
    section("13. Vendor Ledger — PC retention + payment entries")
    db.vendor_ledger.insert_many(pc_ledger_docs)
    ok(f"Inserted {len(pc_ledger_docs)} vendor_ledger entries (RETENTION_HELD + PAYMENT_MADE)")

    # ── 14. Financial State (materialised view) ──────────────────────────────
    section("14. Financial State (derived view)")
    fs_docs = []
    for row in wo_rows:
        code   = row["code"]
        cat_id = cat_id_map.get(code)
        if not cat_id:
            continue
        wo_val  = row["wo_value"]
        pc_vals = [p["pc_value"] for p in pc_rows if p["code"] == code]
        certified = sum(pc_vals)
        balance   = wo_val - wo_val  # remaining_budget is 0 after full commitment
        fs_docs.append({
            "_id":                     oid(),
            "project_id":              project_id,
            "category_id":             cat_id,
            "original_budget":         d128(wo_val),
            "committed_value":         d128(wo_val),
            "certified_value":         d128(certified),
            "balance_budget_remaining": d128(balance),
            "over_commit_flag":        False,
            "last_updated":            now(),
            "version":                 1,
        })
    db.financial_state.insert_many(fs_docs)
    ok(f"Inserted {len(fs_docs)} financial_state records")

    # ── 15. Fund Allocations (OVH only — fund_transfer category) ────────────
    section("15. Fund Allocations (OVH + CSA)")
    fund_alloc_docs = []
    for cat in cats:
        if cat["budget_type"] != "fund_transfer":
            continue
        code   = cat["code"]
        cat_id = cat_id_map.get(code)
        budget_row = ps.get(code, {})
        alloc_orig = budget_row.get("wo_value", Decimal("0"))
        received   = budget_row.get("payment_value", Decimal("0"))
        remaining  = alloc_orig - received
        pc_vals    = [p["pc_value"] for p in pc_rows if p["code"] == code]
        total_exp  = sum(pc_vals)
        cash_hand  = received - total_exp
        fund_alloc_docs.append({
            "_id":                  oid(),
            "project_id":           project_id,
            "category_id":          cat_id,
            "allocation_original":  d128(alloc_orig),
            "allocation_received":  d128(received),
            "allocation_remaining": d128(remaining),
            "cash_in_hand":         d128(cash_hand),
            "total_expenses":       d128(total_exp),
            "last_pc_closed_date":  now(),
            "version":              1,
            "created_at":           now(),
        })
    db.fund_allocations.insert_many(fund_alloc_docs)
    ok(f"Inserted {len(fund_alloc_docs)} fund_allocation records (OVH, CSA)")

    # ── 16. DPRs (one per category from Weekly Progress Report) ─────────────
    section("16. DPRs")
    dpr_docs = []
    ref_date = datetime(2025, 3, 15, 8, 0, 0, tzinfo=timezone.utc)  # reference date
    for code, prog in wp.items():
        pct_val = prog["pct_complete"]
        dpr_docs.append({
            "_id":             oid(),
            "project_id":      project_id,
            "created_by":      str(super_id),
            "date":            ref_date,
            "notes":           (
                f"{code}: {prog['comments']}"
                if prog["comments"]
                else f"{code} — {pct_val:.1f}% complete as of weekly report."
            ),
            "photos":          [],
            "status":          "APPROVED",
            "approved_by":     str(admin_id),
            "approved_at":     ref_date,
            "rejected_by":     None,
            "rejected_at":     None,
            "rejection_reason": None,
            "created_at":      ref_date,
        })
    db.dprs.insert_many(dpr_docs)
    ok(f"Inserted {len(dpr_docs)} DPRs")

    # ── 17. Workers Daily Log + Attendance ───────────────────────────────────
    section("17. Workers Daily Log + Attendance")
    wdl_docs = []
    att_docs = []
    worker_counts = {"CIV": 12, "PLB": 4, "ELC": 4, "FIN": 8,
                     "CRP": 5, "SWP": 2, "LAN": 3, "PRF": 1}

    for code, prog in wp.items():
        count = worker_counts.get(code, 2)
        # Workers Daily Log
        wdl_docs.append({
            "_id":              oid(),
            "organisation_id":  org_id,
            "project_id":       project_id,
            "date":             ref_date.strftime("%Y-%m-%d"),
            "supervisor_id":    str(super_id),
            "supervisor_name":  supervisor_doc["name"],
            "entries": [{
                "vendor_id":       vendor_id,
                "vendor_name":     SEED_VENDOR_NAME,
                "workers_count":   count,
                "skill_type":      code,
                "rate_per_worker": d128(500),
                "remarks":         None,
            }],
            "workers":         [],
            "total_workers":   count,
            "total_hours":     d128(count * 8),
            "weather":         "Sunny",
            "site_conditions": "Good",
            "remarks":         f"{code} — daily log for weekly report cycle",
            "status":          "submitted",
            "created_at":      ref_date,
            "updated_at":      ref_date,
        })

    db.workers_daily_logs.insert_many(wdl_docs)
    ok(f"Inserted {len(wdl_docs)} workers_daily_logs")

    # Attendance — one check-in per supervisor per day
    att_docs.append({
        "_id":               oid(),
        "organisation_id":   org_id,
        "project_id":        project_id,
        "supervisor_id":     str(super_id),
        "date":              ref_date,
        "selfie_url":        None,
        "gps_lat":           15.5988,    # Majorda, Goa approximate GPS
        "gps_lng":           73.7528,
        "check_in_time":     ref_date,
        "verified_by_admin": True,
        "verified_at":       ref_date,
        "verified_user_id":  str(admin_id),
    })
    db.attendance.insert_many(att_docs)
    ok(f"Inserted {len(att_docs)} attendance records")

    # ── Summary ──────────────────────────────────────────────────────────────
    section("SEED COMPLETE — Summary")
    total_wo_val  = sum(r["wo_value"] for r in wo_rows)
    total_pc_val  = sum(r["pc_value"] for r in pc_rows)
    total_pc_ret  = sum(r["retention_val"] for r in pc_rows)
    print(f"  Organisation : {SEED_ORG_NAME}")
    print(f"  Project      : {SEED_PROJECT_NAME}  [{project_id}]")
    print(f"  Categories   : {len(cats)}")
    print(f"  Vendors      : 1 ({SEED_VENDOR_NAME})")
    print(f"  Work Orders  : {len(wo_docs)}  total ₹{total_wo_val:>14,.2f}")
    print(f"  PCs          : {len(pc_docs)}  total ₹{total_pc_val:>14,.2f}  "
          f"retention ₹{total_pc_ret:,.2f}")
    print(f"  DPRs         : {len(dpr_docs)}")
    print(f"  Workers Logs : {len(wdl_docs)}")
    print(f"  Admin login  : amit@thirdangleconcepts.com / {DEFAULT_PASSWORD}")
    print(f"  Super login  : supervisor@thirdangleconcepts.com / {DEFAULT_PASSWORD}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Majorda Villa DB Seed Script")
    parser.add_argument("--cleanup", action="store_true",
                        help="Delete existing seed data first")
    parser.add_argument("--seed", action="store_true",
                        help="Seed data (default if neither flag given)")
    args = parser.parse_args()

    # Default: if no flags given, do both
    do_cleanup = args.cleanup
    do_seed    = args.seed or not args.cleanup

    client = MongoClient(MONGO_URL)
    db     = client[DB_NAME]

    try:
        # Verify connection
        db.command("ping")
        print(f"Connected to MongoDB: {DB_NAME}")
    except Exception as e:
        print(f"FATAL: Cannot connect to MongoDB at {MONGO_URL}\n{e}")
        sys.exit(1)

    if do_cleanup:
        cleanup_db(db)

    if do_seed:
        # Parse Excel
        data = load_excel()

        # Financial validation (PC vs WO)
        validate_pc_vs_wo(data["wo_tracker"], data["pc_tracker"])

        # Seed
        seed(db, data)

    client.close()


if __name__ == "__main__":
    main()
