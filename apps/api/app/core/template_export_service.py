import os
from copy import copy
from datetime import datetime
from io import BytesIO
from typing import Any, Dict

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.core.export_service import ExportService


class TemplateExportService(ExportService):
    """
    Handles exactly formatted Excel and PDF exports based on Live Sheets templates.
    """
    
    TEMPLATE_PATH = os.path.join(
        os.path.dirname(__file__), "..", "..", "templates", "excel", "majorda_templates.xlsx"
    )

    @classmethod
    def get_template_path(cls) -> str:
        if os.path.exists(cls.TEMPLATE_PATH):
            return cls.TEMPLATE_PATH
        # Fallback if run from a different directory (e.g. tests)
        fallback = os.path.join(os.getcwd(), "apps", "api", "templates", "excel", "majorda_templates.xlsx")
        if os.path.exists(fallback):
            return fallback
        return cls.TEMPLATE_PATH

    @classmethod
    def export_work_order_exact(cls, data: Dict[str, Any], fmt: str = "excel") -> bytes:
        """
        Exports a Work Order either directly into the exact Excel Live Sheet
        or to PDF mimicking the exact layout.
        """
        if fmt.lower() == "pdf":
            # Delegate to existing ExportService HTML->PDF pipeline but with specific template
            # For Exact templates, we map line_items to rows for generic fallback
            items = data.get("line_items", data.get("items", []))
            pdf_data = {
                "title": f"Work Order {data.get('wo_ref', '')}",
                "rows": items,
                "metadata": data,
                "company": data.get("company", {})
            }
            # Custom html for Exact WO
            cls.REPORT_TEMPLATES["work_order_exact"] = {
                "template": "work_order_exact.html",
                "columns": [("Sr No", 5), ("Description", 40), ("Qty", 10), ("Rate", 10), ("Total", 15)]
            }
            return cls.export_to_pdf_service("work_order_exact", pdf_data, data.get("company"))
            
        # Excel generation
        wb = openpyxl.load_workbook(cls.get_template_path())
        ws: Worksheet = wb["WO"]
        
        company = data.get("company", {})
        vendor = data.get("vendor", {})
        
        # Fill Headers
        ws["B2"] = vendor.get("name", "Vendor Name")
        ws["B3"] = vendor.get("address_line1", "")
        ws["B4"] = vendor.get("address_line2", "")
        ws["B5"] = vendor.get("address_line3", "")
        ws["C7"] = vendor.get("gst_no", "-")
        
        ws["E4"] = data.get("wo_ref", "")
        wo_date = data.get("wo_date", "")
        if isinstance(wo_date, datetime):
            ws["E5"] = wo_date.strftime("%d-%b-%Y")
        else:
            ws["E5"] = str(wo_date)
            
        ws["E6"] = data.get("code", "")
        ws["E7"] = company.get("name", "Third Angle Concepts (PMC)")
        
        ws["C14"] = vendor.get("contact_person", "")
        ws["D14"] = company.get("contact_person", "")
        
        # Fill Items
        items = data.get("line_items", data.get("items", []))
        start_row = 18
        current_row = start_row
        
        # Standardize items filling
        for idx, item in enumerate(items, 1):
            if current_row >= 44:
                ws.insert_rows(current_row)
                # Copy styling from previous row
                for col in range(2, 7):
                    prev_cell = ws.cell(row=current_row-1, column=col)
                    new_cell = ws.cell(row=current_row, column=col)
                    new_cell.font = copy(prev_cell.font)
                    new_cell.border = copy(prev_cell.border)
                    new_cell.alignment = copy(prev_cell.alignment)
            
            ws.cell(row=current_row, column=2).value = idx # Sr No
            ws.cell(row=current_row, column=3).value = item.get("description", "")
            ws.cell(row=current_row, column=4).value = float(item.get("quantity", 0))
            ws.cell(row=current_row, column=5).value = float(item.get("rate", 0))
            ws.cell(row=current_row, column=6).value = float(item.get("total", 0))
            current_row += 1
            
        # Fill Totals
        offset = max(0, current_row - 44)
        
        ws.cell(row=45 + offset, column=5).value = float(data.get("subtotal", 0))
        ws.cell(row=46 + offset, column=5).value = float(data.get("discount", 0))
        ws.cell(row=47 + offset, column=5).value = float(data.get("total_after_discount", 0))
        ws.cell(row=48 + offset, column=5).value = float(data.get("cgst", 0))
        ws.cell(row=49 + offset, column=5).value = float(data.get("sgst", 0))
        ws.cell(row=50 + offset, column=5).value = float(data.get("grand_total", 0))
        
        # Timeline and warranty
        ws.cell(row=51 + offset, column=3).value = str(data.get("start_date", ""))
        ws.cell(row=52 + offset, column=3).value = str(data.get("end_date", ""))
        ws.cell(row=51 + offset, column=6).value = str(data.get("warranty", ""))
        
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    @classmethod
    def export_payment_certificate_exact(cls, data: Dict[str, Any], fmt: str = "excel") -> bytes:
        """
        Exports a Payment Certificate dynamically matching exact PC Excel Sheet.
        """
        if fmt.lower() == "pdf":
            items = data.get("line_items", data.get("items", []))
            pdf_data = {
                "title": f"Payment Certificate {data.get('pc_ref', '')}",
                "rows": items,
                "metadata": data,
                "company": data.get("company", {})
            }
            cls.REPORT_TEMPLATES["payment_certificate_exact"] = {
                "template": "payment_certificate_exact.html",
                "columns": [("Sr No", 5), ("Description", 40), ("Rate", 10), ("Qty", 10), ("Unit", 10), ("Total", 15)]
            }
            return cls.export_to_pdf_service("payment_certificate_exact", pdf_data, data.get("company"))
            
        wb = openpyxl.load_workbook(cls.get_template_path())
        ws: Worksheet = wb["PC"]
        
        vendor = data.get("vendor", {})
        
        ws["B2"] = vendor.get("name", "Vendor Name")
        ws["G5"] = data.get("pc_ref", "")
        
        pc_date = data.get("pc_date", "")
        ws["G6"] = pc_date.strftime("%d-%b-%Y") if isinstance(pc_date, datetime) else str(pc_date)
        
        ws["C7"] = vendor.get("gst_no", "-")
        ws["G7"] = data.get("wo_ref", "")
        
        # PC Items typically start at row 10 (Sr No is row 9)
        items = data.get("line_items", data.get("line_items", []))
        start_row = 10
        current_row = start_row
        
        for idx, item in enumerate(items, 1):
            if current_row >= 43:
                ws.insert_rows(current_row)
                for col in range(2, 8):
                    prev_cell = ws.cell(row=current_row-1, column=col)
                    new_cell = ws.cell(row=current_row, column=col)
                    new_cell.font = copy(prev_cell.font)
                    new_cell.border = copy(prev_cell.border)
                    new_cell.alignment = copy(prev_cell.alignment)
            
            ws.cell(row=current_row, column=2).value = idx # Sr. No.
            ws.cell(row=current_row, column=3).value = item.get("description", "") # Scope of Work
            ws.cell(row=current_row, column=4).value = float(item.get("rate", 0))
            ws.cell(row=current_row, column=5).value = float(item.get("quantity", 0))
            ws.cell(row=current_row, column=6).value = str(item.get("unit", ""))
            ws.cell(row=current_row, column=7).value = float(item.get("total", 0))
            current_row += 1
            
        offset = max(0, current_row - 43)
        
        # Inject totals
        ws.cell(row=44 + offset, column=3).value = str(data.get("pmc_comments", ""))
        
        ws.cell(row=44 + offset, column=7).value = float(data.get("subtotal", 0))
        ws.cell(row=45 + offset, column=7).value = float(data.get("retention", 0))
        ws.cell(row=46 + offset, column=7).value = float(data.get("total_after_retention", 0))
        ws.cell(row=47 + offset, column=7).value = float(data.get("cgst", 0))
        ws.cell(row=48 + offset, column=7).value = float(data.get("sgst", 0))
        ws.cell(row=49 + offset, column=7).value = float(data.get("grand_total", 0))
        
        out = BytesIO()
        wb.save(out)
        return out.getvalue()
