import logging
import tracemalloc
from app.core.time import now

from io import BytesIO
from typing import Any, Dict, Optional, List
import io

logger = logging.getLogger(__name__)


class ExportService:

    """
    Sovereign Export Engine (Ported from Legacy Core).
    Handles Excel and PDF templates for construction reports.
    """

    REPORT_TEMPLATES = {
        "project_summary": {
            "description": "Project-level financial summary",
            "columns": [
                ("CODE", 10),
                ("Description", 30),
                ("Budget", 12),
                ("Committed", 12),
                ("Certified", 12),
                ("Remaining", 12),
                ("Status", 10),
            ],
            "template": "project_summary.html"
        },
        "work_order_tracker": {
            "description": "Work Order tracking report",
            "columns": [
                ("CODE", 10),
                ("WO Reference", 15),
                ("Vendor", 20),
                ("WO Value", 15),
                ("Retention", 15),
                ("Date", 12),
                ("Status", 13),
            ],
            "template": "work_order_tracker.html"
        },
        "payment_certificate_tracker": {
            "description": "PC tracking report",
            "columns": [
                ("CODE", 10),
                ("PC Reference", 15),
                ("Vendor", 20),
                ("Total Value", 15),
                ("Date", 12),
                ("Certified", 15),
                ("Status", 13),
            ],
            "template": "payment_certificate_tracker.html"
        },
        "petty_cash_tracker": {
            "description": "Petty Cash tracking report",
            "columns": [
                ("Date", 15),
                ("Ref", 20),
                ("Amount", 20),
                ("Description", 45),
            ],
            "template": "petty_cash_tracker.html"
        },
        "csa_report": {
            "description": "CSA Report",
            "columns": [
                ("CODE", 10),
                ("Ref", 15),
                ("Description", 40),
                ("Qty", 10),
                ("Date", 12),
            ],
            "template": "csa_report.html"
        },
        "attendance": {
            "description": "Attendance tracking",
            "columns": [
                ("Date", 12),
                ("Worker Name", 25),
                ("Category", 15),
                ("Check In", 12),
                ("Check Out", 12),
            ],
            "template": "attendance.html"
        },
        "dpr_report": {
            "description": "Daily Progress Report",
            "columns": [],
            "template": "dpr_report.html"
        },
        "weekly_progress": {
            "description": "Weekly Progress",
            "columns": [
                ("CODE", 10),
                ("Ref", 15),
                ("Vendor", 20),
                ("Progress (%)", 15),
                ("Status", 15),
            ],
            "template": "progress_report.html"
        },
        "15_days_progress": {
            "description": "15-Day Progress",
            "columns": [
                ("CODE", 10),
                ("Ref", 15),
                ("Vendor", 20),
                ("Progress (%)", 15),
                ("Status", 15),
            ],
            "template": "progress_report.html"
        },
        "monthly_progress": {
            "description": "Monthly Progress",
            "columns": [
                ("CODE", 10),
                ("Ref", 15),
                ("Vendor", 20),
                ("Progress (%)", 15),
                ("Status", 15),
            ],
            "template": "progress_report.html"
        },
        "scheduler_gantt": {
            "description": "High-fidelity Gantt Chart Export",
            "columns": [],
            "template": "scheduler_gantt.html"
        },
    }

    @staticmethod
    def format_currency(value: Any) -> str:
        if value is None:
            return "₹0.00"
        try:
            val = float(value)
            import math
            if math.isnan(val) or math.isinf(val):
                return "₹0.00"
            
            is_negative = val < 0
            val_abs = abs(val)
            
            s = f"{val_abs:.2f}"
            parts = s.split('.')
            integer_part = parts[0]
            decimal_part = parts[1]
            
            if len(integer_part) <= 3:
                formatted_integer = integer_part
            else:
                last_three = integer_part[-3:]
                remaining = integer_part[:-3]
                groups = []
                while len(remaining) > 2:
                    groups.append(remaining[-2:])
                    remaining = remaining[:-2]
                if remaining:
                    groups.append(remaining)
                groups.reverse()
                formatted_integer = ",".join(groups) + "," + last_three
                
            return f"₹{'-' if is_negative else ''}{formatted_integer}.{decimal_part}"
        except Exception:
            return str(value)

    @staticmethod
    def validate_report_type(rt: str) -> bool:
        return rt in ExportService.REPORT_TEMPLATES

    @staticmethod
    def export_to_excel(
        report_type: str,
        report_data: Dict[str, Any],
        company_info: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        # Freeze the first row
        ws.freeze_panes = "A2"

        config = ExportService.REPORT_TEMPLATES.get(report_type, {})
        columns = config.get("columns", [])

        # Visual style helpers for Luxury Industrial theme
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        
        data_font = Font(name="Segoe UI", size=10)
        total_font = Font(name="Segoe UI", size=10, bold=True)
        
        thin_side = Side(border_style="thin", color="CBD5E1")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        double_bottom = Border(top=Side(style="thin", color="CBD5E1"), bottom=Side(style="double", color="0F172A"))

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        # Detect currency columns
        currency_cols = set()
        for col_idx, (col_name, _) in enumerate(columns, 1):
            c_name = col_name.lower()
            if any(k in c_name for k in ["budget", "committed", "certified", "remaining", "value", "retention", "amount", "total"]):
                if not any(k in c_name for k in ["status", "code", "ref", "date"]):
                    currency_cols.add(col_idx)

        # Header logic
        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = align_center if col_idx in currency_cols or "date" in col_name.lower() else align_left
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Helper to parse formatted currency strings back to numeric float
        def parse_currency(val: Any) -> Optional[float]:
            if val is None:
                return None
            s = str(val).strip()
            if not s:
                return None
            # Check if it has a currency indicator
            if "₹" in s or "$" in s or s.startswith("-") or s.replace(".", "").replace(",", "").isdigit():
                clean = s.replace("₹", "").replace("$", "").replace(",", "").replace(" ", "").strip()
                try:
                    return float(clean)
                except ValueError:
                    return None
            return None

        # Data logic
        rows = report_data.get("rows", [])
        last_row = 1
        for row_idx, row_data in enumerate(rows, 2):
            last_row = row_idx
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Check if it is a currency column or parses as currency
                parsed_float = parse_currency(value)
                if parsed_float is not None and (col_idx in currency_cols or (isinstance(value, str) and "₹" in value)):
                    cell.value = parsed_float
                    cell.number_format = '[$₹-4009] #,##,##0.00'
                    cell.alignment = align_right
                    currency_cols.add(col_idx)  # Ensure this column is recognized as currency
                else:
                    cell.value = str(value) if value is not None else ""
                    if col_idx in currency_cols:
                        cell.alignment = align_right
                    elif "date" in columns[col_idx-1][0].lower():
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_left

        # Inject dynamic =SUM() formulas in the totals row
        if rows:
            totals_row_idx = last_row + 1
            # Column 1 gets "Total"
            cell = ws.cell(row=totals_row_idx, column=1, value="Total")
            cell.font = total_font
            cell.border = double_bottom
            cell.alignment = align_left

            # Fill in the rest of the columns
            for col_idx in range(2, len(columns) + 1):
                cell = ws.cell(row=totals_row_idx, column=col_idx)
                cell.font = total_font
                cell.border = double_bottom
                if col_idx in currency_cols:
                    col_letter = get_column_letter(col_idx)
                    cell.value = f"=SUM({col_letter}2:{col_letter}{last_row})"
                    cell.number_format = '[$₹-4009] #,##,##0.00'
                    cell.alignment = align_right
                else:
                    cell.value = ""

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    @staticmethod
    def export_to_pdf_service(
        report_type: str,
        report_data: Dict[str, Any],
        company_info: Optional[Dict[str, Any]] = None,
    ) -> bytes:
        import os
        try:
            import jinja2
        except ImportError:
            logger.error("Jinja2 package is not installed. Please install it using 'pip install jinja2'.")
            raise

        # Setup Jinja2 environment
        search_paths = [
            os.path.join(os.getcwd(), "templates"),
            os.path.join(os.path.dirname(__file__), "..", "..", "..", "templates"),
            os.path.join(os.path.dirname(__file__), "..", "templates"),
            "app/core/templates",
            "templates"
        ]

        # BUG-018: Memory Profiling Instrumentation (Dev Mode)
        tracemalloc.start()
        snapshot_start = tracemalloc.take_snapshot()

        template_dir = next((p for p in search_paths if os.path.exists(p)), search_paths[0])
        loader = jinja2.FileSystemLoader(template_dir)
        env = jinja2.Environment(loader=loader)

        config = ExportService.REPORT_TEMPLATES.get(report_type, {})
        template_name = config.get("template", "generic_report.html")

        try:
            template = env.get_template(template_name)
        except Exception as te:
            logger.warning(f"Template {template_name} not found in {template_dir}: {te}. Using string fallback.")
            template = env.from_string("""
                <html>
                <body>
                    <h1>{{ report_title }}</h1>
                    <table border="1">
                        <thead>
                            <tr>
                                {% for col in columns %}
                                <th>{{ col[0] }}</th>
                                {% endfor %}
                            </tr>
                        </thead>
                        <tbody>
                            {% for row in rows %}
                            <tr>
                                {% for cell in row %}
                                <td>{{ cell }}</td>
                                {% endfor %}
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </body>
                </html>
            """)

        # BUG-018: PDF row limit guard
        MAX_PDF_ROWS = 500
        rows = report_data.get("rows", [])
        original_count = len(rows)
        if original_count > MAX_PDF_ROWS:
            logger.warning(f"Report has {original_count} rows, truncating to {MAX_PDF_ROWS}")
            report_data["rows"] = rows[:MAX_PDF_ROWS]
            if "metadata" not in report_data:
                report_data["metadata"] = {}
            report_data["metadata"]["truncated"] = True
            report_data["metadata"]["original_count"] = original_count

        # Build context — explicit keys take priority; extra keys from report_data
        # (e.g. tasks, months for specialized templates) are merged in without
        # duplicating the named keys (which would cause a TypeError).
        context: dict = {
            "report_title": report_data.get("title", report_type.replace("_", " ").title()),
            "rows": report_data.get("rows", []),
            "columns": config.get("columns", []),
            "totals": report_data.get("totals", {}),
            "metadata": report_data.get("metadata", {}),
            "company": company_info or {"name": "TAC PMC", "address": "Sovereign HQ"},
            "now": now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        for k, v in report_data.items():
            if k not in context:
                context[k] = v

        # BUG-018: Performance optimization
        # For large reports (>100 rows), skip WeasyPrint (which renders full HTML/CSS)
        # and use ReportLab's SimpleDocTemplate which handles memory/pagination better.
        if len(report_data.get("rows", [])) > 100:
            logger.info(f"Report has {len(report_data.get('rows', []))} rows. Using ReportLab primary engine.")
            pdf_bytes = ExportService._generate_pdf_reportlab(report_data, config, report_type)
        else:
            html_out = template.render(**context)
            try:
                # Check if WeasyPrint is likely to work before attempting import
                # This avoids system-level DLL hangs on Windows if missing GTK
                from weasyprint import HTML
                pdf_bytes = HTML(string=html_out).write_pdf()
            except (ImportError, OSError, Exception) as e:
                # Fallback to ReportLab if WeasyPrint system dependencies are missing
                logger.error(
                    f"WeasyPrint failed ({type(e).__name__}: {e}), falling back to ReportLab. "
                    f"HTML length: {len(html_out)}"
                )
                try:
                    pdf_bytes = ExportService._generate_pdf_reportlab(report_data, config, report_type)
                except Exception as re:
                    logger.error(f"ReportLab fallback also failed: {re}")
                    # Cleanup tracemalloc before raising
                    tracemalloc.stop()
                    raise RuntimeError(
                        "PDF generation failed: Both WeasyPrint and ReportLab engines are "
                        "unavailable or encountered a terminal error."
                    )

        # Log memory metrics
        snapshot_end = tracemalloc.take_snapshot()
        top_stats = snapshot_end.compare_to(snapshot_start, 'lineno')
        peak = tracemalloc.get_traced_memory()[1] / (1024 * 1024)
        logger.info(f"PDF Export Memory Peak: {peak:.2f} MB | Rows: {len(report_data.get('rows', []))}")
        tracemalloc.stop()

        return pdf_bytes

    @staticmethod
    def _generate_pdf_reportlab(report_data: Dict[str, Any], config: Dict[str, Any], report_type: Optional[str] = None) -> bytes:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from reportlab.platypus import HRFlowable

        # Determine dynamic orientation based on report type or title matching
        if not report_type:
            title_lower = str(report_data.get("title", "")).lower()
            if "work order" in title_lower:
                report_type = "work_order_tracker"
            elif "payment certificate" in title_lower:
                report_type = "payment_certificate_tracker"

        is_landscape = report_type in ["work_order_tracker", "payment_certificate_tracker"]
        page_size = landscape(A4) if is_landscape else A4
        
        # Sizing proportional layout width (Content area: A4 dimensions - 60 points for margins)
        content_width = 781.89 if is_landscape else 535.27

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        elements = []
        styles = getSampleStyleSheet()

        # Add custom styles for Luxury Industrial aesthetic
        if "Title" not in styles:
            styles.add(ParagraphStyle(name="Title", fontSize=18, leading=22, alignment=TA_CENTER, spaceAfter=20, fontName="Helvetica-Bold", textColor=colors.HexColor("#0f172a")))
        
        # Luxury Header Style
        header_style = ParagraphStyle(
            name="LuxuryHeader",
            parent=styles["Normal"],
            fontSize=10,
            leading=12,
            textColor=colors.HexColor("#64748b"),
            fontName="Helvetica"
        )

        # Luxury Colors
        GOLD_ACCENT = colors.HexColor("#b45309")  # Muted Amber/Gold
        CHARCOAL = colors.HexColor("#0f172a")

        # Branded Two-Column Header
        logo_flowable = None
        logo_paths = [
            "d:/_repos/TAC-PMC-CRM/apps/mobile/assets/images/logo.png",
            "apps/mobile/assets/images/logo.png",
            "../mobile/assets/images/logo.png"
        ]
        import os
        for p in logo_paths:
            if os.path.exists(p):
                try:
                    from reportlab.platypus import Image as RLImage
                    logo_flowable = RLImage(p, width=32, height=32)
                    break
                except Exception:
                    pass
        
        if not logo_flowable:
            from reportlab.graphics.shapes import Drawing, Circle, String
            logo_flowable = Drawing(32, 32)
            logo_flowable.add(Circle(16, 16, 16, fillColor=GOLD_ACCENT, strokeColor=None))
            logo_flowable.add(String(11, 9, "T", fontName="Helvetica-Bold", fontSize=20, fillColor=colors.white))

        company = report_data.get("company") or {"name": "Third Angle Concepts (PMC)", "address": "Sovereign HQ, India"}
        company_name = company.get("name", "Third Angle Concepts (PMC)")
        company_addr = company.get("address", "Sovereign HQ, India")

        left_paragraphs = [
            logo_flowable,
            Spacer(1, 4),
            Paragraph(f"<b>{company_name.upper()}</b>", ParagraphStyle(
                name="CompName", parent=styles["Normal"], fontSize=11, leading=13, textColor=CHARCOAL, fontName="Helvetica-Bold"
            )),
            Paragraph(company_addr, ParagraphStyle(
                name="CompAddr", parent=styles["Normal"], fontSize=8, leading=10, textColor=colors.HexColor("#64748b")
            ))
        ]

        project_name = report_data.get("project_name", "Unknown Project")
        project_code = report_data.get("project_code", "N/A")
        gen_name = report_data.get("generator_name", "System User")
        gen_role = report_data.get("generator_role", "Administrator")
        gen_date = now().strftime('%d-%b-%Y %H:%M')

        report_title = report_data.get("title", "Project Report")

        right_paragraphs = [
            Paragraph(report_title.upper(), ParagraphStyle(
                name="RepTitle", parent=styles["Normal"], fontSize=14, leading=16, textColor=GOLD_ACCENT, fontName="Helvetica-Bold", alignment=TA_RIGHT
            )),
            Spacer(1, 4),
            Paragraph(f"<b>Project:</b> {project_name} ({project_code})", ParagraphStyle(
                name="ProjInfo", parent=styles["Normal"], fontSize=9, leading=12, textColor=CHARCOAL, alignment=TA_RIGHT
            )),
            Paragraph(f"<b>Generated by:</b> {gen_name} ({gen_role})", ParagraphStyle(
                name="GenInfo", parent=styles["Normal"], fontSize=8, leading=11, textColor=colors.HexColor("#64748b"), alignment=TA_RIGHT
            )),
            Paragraph(f"<b>Date:</b> {gen_date}", ParagraphStyle(
                name="GenDate", parent=styles["Normal"], fontSize=8, leading=11, textColor=colors.HexColor("#64748b"), alignment=TA_RIGHT
            ))
        ]

        header_table_data = [[left_paragraphs, right_paragraphs]]
        header_table = Table(header_table_data, colWidths=[content_width * 0.5, content_width * 0.5])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(header_table)
        elements.append(HRFlowable(width="100%", thickness=1.5, color=GOLD_ACCENT, spaceBefore=10, spaceAfter=15))

        # Truncation Warning
        if report_data.get("metadata", {}).get("truncated"):
            warn_style = ParagraphStyle(name="Warning", fontSize=9, textColor=colors.red, fontName="Helvetica-Bold")
            orig = report_data["metadata"].get("original_count", 500)
            elements.append(Paragraph(f"⚠️ NOTICE: Report truncated to 500 rows (Original: {orig} rows) to preserve system memory.", warn_style))
            elements.append(Spacer(1, 10))

        # Table Data
        headers = [col[0] for col in config.get("columns", [])]
        col_widths_raw = [col[1] for col in config.get("columns", [])]

        total_relative = sum(col_widths_raw)
        if total_relative > 0:
            col_widths = [(w / total_relative) * content_width for w in col_widths_raw]
        else:
            col_widths = [content_width / max(1, len(headers))] * max(1, len(headers))

        rows = report_data.get("rows", [])
        
        # Format cells
        def wrap_cell(val, idx):
            s_val = str(val) if val is not None else ""
            style = ParagraphStyle(name=f"Cell_Style_{idx}_{now().microsecond}", parent=styles["Normal"], fontSize=8, leading=10)
            # Align currency to right if header contains budget/amount/total etc.
            if headers and idx < len(headers):
                h = headers[idx].lower()
                if any(k in h for k in ["budget", "amount", "total", "certified", "committed", "remaining", "value", "retention"]):
                    style = ParagraphStyle(name=f"Cell_Right_{idx}_{now().microsecond}", parent=style, alignment=TA_RIGHT)
            return Paragraph(s_val, style)

        data = [[Paragraph(str(h), ParagraphStyle(name=f"H_{i}", parent=styles["Normal"], fontSize=9, fontName="Helvetica-Bold", textColor=colors.whitesmoke)) for i, h in enumerate(headers)]]
        for row in rows:
            if isinstance(row, dict):
                formatted_row = [wrap_cell(row.get(h, ""), i) for i, h in enumerate(headers)]
            else:
                formatted_row = [wrap_cell(cell, i) for i, cell in enumerate(row)]
            data.append(formatted_row)

        # Create Table
        if data:
            t = Table(data, colWidths=col_widths, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), CHARCOAL),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
            ]))
            elements.append(t)

        # Totals Section (if available)
        totals = report_data.get("totals", {})
        if totals:
            elements.append(Spacer(1, 20))
            elements.append(Paragraph("SUMMARY TOTALS", ParagraphStyle(name="TotalHeader", parent=styles["Heading3"], fontSize=10, textColor=GOLD_ACCENT, spaceAfter=8)))
            total_data = [[k, ExportService.format_currency(v)] for k, v in totals.items()]
            tt = Table(total_data, colWidths=[content_width * 0.4, content_width * 0.3])
            tt.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('LINEABOVE', (0, 0), (-1, 0), 1.5, CHARCOAL),
                ('TEXTCOLOR', (0, 0), (-1, -1), CHARCOAL),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(tt)

        doc.build(elements, canvasmaker=NumberedCanvas)
        return buffer.getvalue()

    @staticmethod
    def merge_pdfs(base_pdf: bytes, attachments: List[bytes]) -> bytes:
        """
        Merges multiple PDF documents into a single PDF.
        The base_pdf is placed first, followed by each attachment in order.
        Adds page numbering (e.g. "Page 1 / 10") to the bottom of each page.
        """
        from pypdf import PdfWriter, PdfReader
        from reportlab.pdfgen import canvas

        writer = PdfWriter()

        # 1. Add all pages to the writer
        # Add base PDF
        try:
            reader = PdfReader(io.BytesIO(base_pdf))
            for page in reader.pages:
                writer.add_page(page)
        except Exception as e:
            logger.error(f"Failed to read base PDF for merging: {e}")
            return base_pdf

        # Add attachments
        for i, attachment_bytes in enumerate(attachments):
            try:
                reader = PdfReader(io.BytesIO(attachment_bytes))
                for page in reader.pages:
                    writer.add_page(page)
            except Exception as e:
                logger.error(f"Failed to append attachment {i} during PDF merge: {e}")
                continue

        total_pages = len(writer.pages)
        if total_pages == 0:
            return base_pdf

        # 2. Add page numbers by overlaying a reportlab-generated PDF
        output_writer = PdfWriter()

        for i in range(total_pages):
            page = writer.pages[i]

            # Create a temporary PDF with just the page number
            packet = io.BytesIO()
            can = canvas.Canvas(packet, pagesize=(page.mediabox.width, page.mediabox.height))

            # Set font and size
            can.setFont("Helvetica", 9)
            can.setFillColorRGB(0.5, 0.5, 0.5)  # Grey color

            # Draw page number at bottom center
            page_num_text = f"Page {i + 1} of {total_pages}"
            text_width = can.stringWidth(page_num_text, "Helvetica", 9)

            # Position: 20 units from bottom, centered
            can.drawString((float(page.mediabox.width) - text_width) / 2, 20, page_num_text)
            can.save()

            packet.seek(0)
            overlay_reader = PdfReader(packet)
            overlay_page = overlay_reader.pages[0]

            # Merge overlay onto the original page
            page.merge_page(overlay_page)
            output_writer.add_page(page)

        out = io.BytesIO()
        output_writer.write(out)
        return out.getvalue()


class NumberedCanvas(object):
    """A ReportLab canvas that dynamically tracks total page count and prints page numbers."""
    def __new__(cls, *args, **kwargs):
        from reportlab.pdfgen import canvas
        
        class _NumberedCanvasInner(canvas.Canvas):
            def __init__(self, *args, **kwargs):
                super().__init__(*args, **kwargs)
                self._saved_page_states = []

            def showPage(self):
                self._saved_page_states.append(dict(self.__dict__))
                self._startPage()

            def save(self):
                num_pages = len(self._saved_page_states)
                for state in self._saved_page_states:
                    self.__dict__.update(state)
                    self.draw_page_number(num_pages)
                    super().showPage()
                super().save()

            def draw_page_number(self, page_count):
                self.saveState()
                self.setFont("Helvetica", 8)
                from reportlab.lib import colors
                self.setFillColor(colors.HexColor("#64748b"))
                
                width, height = self._pagesize
                
                text = f"Page {self._pageNumber} of {page_count}"
                text_width = self.stringWidth(text, "Helvetica", 8)
                
                # Draw dynamic footer
                self.drawString((width - text_width) / 2, 20, text)
                
                # Thin divider line above footer
                self.setStrokeColor(colors.HexColor("#e2e8f0"))
                self.setLineWidth(0.5)
                self.line(30, 32, width - 30, 32)
                
                self.drawString(30, 20, "Confidential - Third Angle Concepts (PMC)")
                self.restoreState()
                
        return _NumberedCanvasInner(*args, **kwargs)
